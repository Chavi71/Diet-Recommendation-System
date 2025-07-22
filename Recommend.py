from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
from PIL import Image, ImageTk, ImageFilter

ENTRY1 = ENTRY2 = ENTRY3 = ENTRY4 = None
RADIO_VAR = GYM_VAR = LIFESTYLE_VAR = None
DAY_VAR = MONTH_VAR = YEAR_VAR = None
BMI_LABEL = None
bg_photo = None

food_recommendations = {
    "Very severely underweight": "Consume calorie-dense foods like nuts, dried fruits, cheese, and healthy fats like avocado and olive oil.",
    "Severely underweight": "Increase intake of protein-rich foods like lean meats, fish, eggs, and legumes.",
    "Underweight": "Include more whole grains, dairy products, and starchy vegetables in your diet.",
    "Normal": "Maintain a balanced diet with a variety of fruits, vegetables, whole grains, and lean proteins.",
    "Overweight": "Focus on eating more vegetables, fruits, whole grains, and lean proteins while reducing intake of high-calorie foods.",
    "Moderately obese": "Reduce portion sizes, avoid sugary drinks, and incorporate more fiber-rich foods like vegetables and legumes.",
    "Severely obese": "Limit intake of processed and high-fat foods, and prioritize a diet rich in vegetables, lean proteins, and whole grains.",
    "Super obese": "Consult with a healthcare provider for a personalized diet plan. Focus on nutrient-dense foods and avoid high-calorie, low-nutrient items."
}


def get_name(): return ENTRY1.get()
def get_age(): return int(ENTRY2.get())
def get_gender(): return RADIO_VAR.get()
def get_weight(): return float(ENTRY3.get())
def get_height(): return float(ENTRY4.get())
def get_gym(): return GYM_VAR.get()
def get_lifestyle(): return LIFESTYLE_VAR.get()
def get_dob(): return f"{DAY_VAR.get()}-{MONTH_VAR.get()}-{YEAR_VAR.get()}"


def calculate_age():
    try:
        day = int(DAY_VAR.get())
        month = MONTHS.index(MONTH_VAR.get()) + 1
        year = int(YEAR_VAR.get())
        dob = datetime(year, month, day)
        today = datetime.today()
        age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        ENTRY2.delete(0, END)
        ENTRY2.insert(0, str(age))
    except ValueError:
        messagebox.showinfo("Error", "Invalid date of birth")


def save_to_excel(name, age, gender, dob, height, weight, bmi, remark, gym, lifestyle):
    filename = "bmi_data.xlsx"
    try:
        if os.path.exists(filename):
            workbook = load_workbook(filename)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Name", "Age", "Gender", "Date of Birth", "Height (cm)", "Weight (kg)", "BMI", "Remarks", "Gym", "Lifestyle"])
        sheet.append([name, age, gender, dob, height, weight, bmi, remark, gym, lifestyle])
        workbook.save(filename)
        messagebox.showinfo("Success", "Data saved to Excel file successfully!")
    except Exception as e:
        messagebox.showinfo("Error", f"Error saving data to Excel: {e}")


def calculate_bmi(event=None):
    try:
        name = get_name()
        age = get_age()
        gender = get_gender()
        dob = get_dob()
        height = get_height()
        weight = get_weight()
        gym = get_gym()
        lifestyle = get_lifestyle()

        height_m = height / 100.0
        bmi = weight / (height_m ** 2)
    except ZeroDivisionError:
        messagebox.showinfo("Result", "Please enter positive height!!")
    except ValueError:
        messagebox.showinfo("Result", "Please enter valid data!")
    else:
        if bmi <= 15.0:
            remark = "Very severely underweight"
        elif 15.0 < bmi <= 16.0:
            remark = "Severely underweight"
        elif 16.0 < bmi < 18.5:
            remark = "Underweight"
        elif 18.5 <= bmi <= 25.0:
            remark = "Normal"
        elif 25.0 < bmi <= 30:
            remark = "Overweight"
        elif 30.0 < bmi <= 35.0:
            remark = "Moderately obese"
        elif 35.0 < bmi <= 40.0:
            remark = "Severely obese"
        else:
            remark = "Super obese"

        recommendation = food_recommendations.get(remark, "Consult with a healthcare provider for personalized advice.")
        res = f"Your BMI is {bmi:.2f}\nRemarks: {remark}\nGym: {gym}\nLifestyle: {lifestyle}\nFood Recommendations: {recommendation}"
        messagebox.showinfo("Result", res)
        BMI_LABEL.config(text=res)
        save_to_excel(name, age, gender, dob, height, weight, bmi, remark, gym, lifestyle)


if __name__ == '__main__':
    TOP = Tk()
    TOP.bind("<Return>", calculate_bmi)
    TOP.geometry("500x700")
    TOP.title("BMI Calculator")
    TOP.resizable(True, True)

    try:
        bg_image = Image.open("pic.jpg")
        bg_image = bg_image.resize((1280, 1280), Image.LANCZOS)
        blurred_bg = bg_image.filter(ImageFilter.BLUR)
        bg_photo = ImageTk.PhotoImage(blurred_bg)
    except Exception as e:
        messagebox.showinfo("Error", f"Unable to load background image: {e}")

    bg_label = Label(TOP, image=bg_photo)
    bg_label.place(relwidth=1, relheight=1)

    FORM_FRAME = Frame(TOP, bg="#ffc0cb", bd=0, relief=SUNKEN)
    FORM_FRAME.place(relx=0.5, rely=0.5, anchor=CENTER)

    Label(FORM_FRAME, text="DIET RECOMMENDATION SYSTEM", bg="#ffc0cb", font=("Helvetica", 20, "bold"), pady=10).grid(row=0, columnspan=2)

    Label(FORM_FRAME, text="Enter Name:", bg="#ffc0cb", font=("Helvetica", 12, "bold"), bd=6).grid(row=1, column=0, sticky=W, padx=10, pady=5)
    ENTRY1 = Entry(FORM_FRAME, bd=8, width=30, font="Roboto 12")
    ENTRY1.grid(row=1, column=1, pady=5)

    Label(FORM_FRAME, text="Select Date of Birth:", bg="#ffc0cb", font=("Helvetica", 12, "bold"), bd=6).grid(row=2, column=0, sticky=W, padx=10, pady=5)
    DOB_FRAME = Frame(FORM_FRAME, bg="#ffc0cb")
    DOB_FRAME.grid(row=2, column=1, pady=5)

    DAY_VAR = StringVar(value="1")
    OptionMenu(DOB_FRAME, DAY_VAR, *[str(i) for i in range(1, 32)], command=lambda _: calculate_age()).grid(row=0, column=0, padx=2)

    MONTH_VAR = StringVar(value="Jan")
    MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    OptionMenu(DOB_FRAME, MONTH_VAR, *MONTHS, command=lambda _: calculate_age()).grid(row=0, column=1, padx=2)

    YEAR_VAR = StringVar(value="2000")
    OptionMenu(DOB_FRAME, YEAR_VAR, *[str(i) for i in range(1970, 2030)], command=lambda _: calculate_age()).grid(row=0, column=2, padx=2)

    Label(FORM_FRAME, text="Enter Age:", bg="#ffc0cb", font=("Helvetica", 12, "bold"), bd=6).grid(row=3, column=0, sticky=W, padx=10, pady=5)
    ENTRY2 = Entry(FORM_FRAME, bd=8, width=8, font="Roboto 12")
    ENTRY2.grid(row=3, column=1, pady=5)

    Label(FORM_FRAME, text="Enter Gender:", bg="#ffc0cb", font=("Helvetica", 12, "bold"), bd=6).grid(row=4, column=0, sticky=W, padx=10, pady=5)
    RADIO_VAR = StringVar(value="Male")
    Radiobutton(FORM_FRAME, text="Male", variable=RADIO_VAR, value="Male", bg="#ffc0cb", font=("Helvetica", 12)).grid(row=4, column=1, sticky=W)
    Radiobutton(FORM_FRAME, text="Female", variable=RADIO_VAR, value="Female", bg="#ffc0cb", font=("Helvetica", 12)).grid(row=4, column=1)
    Radiobutton(FORM_FRAME, text="Other", variable=RADIO_VAR, value="Other", bg="#ffc0cb", font=("Helvetica", 12)).grid(row=4, column=1, sticky=E)

    Label(FORM_FRAME, text="Enter Weight (in kg):", bg="#ffc0cb", font=("Helvetica", 12, "bold"), bd=6).grid(row=5, column=0, sticky=W, padx=10, pady=5)
    ENTRY3 = Entry(FORM_FRAME, bd=8, width=8, font="Roboto 12")
    ENTRY3.grid(row=5, column=1, pady=5)

    Label(FORM_FRAME, text="Enter Height (in cm):", bg="#ffc0cb", font=("Helvetica", 12, "bold"), bd=6).grid(row=6, column=0, sticky=W, padx=10, pady=5)
    ENTRY4 = Entry(FORM_FRAME, bd=8, width=8, font="Roboto 12")
    ENTRY4.grid(row=6, column=1, pady=5)

    Label(FORM_FRAME, text="Do you go to the gym?", bg="#ffc0cb", font=("Helvetica", 12, "bold"), bd=6).grid(row=7, column=0, sticky=W, padx=10, pady=5)
    GYM_VAR = StringVar(value="No")
    Radiobutton(FORM_FRAME, text="Yes", variable=GYM_VAR, value="Yes", bg="#ffc0cb", font=("Helvetica", 12)).grid(row=7, column=1, sticky=W)
    Radiobutton(FORM_FRAME, text="No", variable=GYM_VAR, value="No", bg="#ffc0cb", font=("Helvetica", 12)).grid(row=7, column=1)

    Label(FORM_FRAME, text="Select Lifestyle:", bg="#ffc0cb", font=("Helvetica", 12, "bold"), bd=6).grid(row=8, column=0, sticky=W, padx=10, pady=5)
    LIFESTYLE_VAR = StringVar(value="Sedentary")
    OptionMenu(FORM_FRAME, LIFESTYLE_VAR, "Sedentary", "Active").grid(row=8, column=1, pady=5)

    Button(FORM_FRAME, bg="#2187e7", bd=12, text="Calculate BMI", padx=33, pady=15,
           command=calculate_bmi, font=("Helvetica", 20, "bold")).grid(row=9, columnspan=2, pady=10)

    BMI_LABEL = Label(FORM_FRAME, bg="#ffc0cb", text="", font=("Helvetica", 12, "bold"), pady=10)
    BMI_LABEL.grid(row=10, columnspan=2, pady=10)

    TOP.mainloop()
